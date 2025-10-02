import io
import os
import re
import unicodedata
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import column_index_from_string, get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# =========================
# 고정 설정
# =========================
DATA_DIR = "./data"
SHEET_NAME = 0             # 첫 시트
HEADER_ROW_INDEX = 1       # A2:AG2 헤더 → header=1 (0-base)
USECOLS_RANGE = "A:AG"     # 읽기 열 구간
TARGET_SUM_ROW = 174       # 합계 기준 행

# 반/번호/이름 자동 탐지 후보
CANDIDATE_CLASS_COLS = ["반", "학급", "Class"]
CANDIDATE_NO_COLS    = ["번호", "번", "No"]
CANDIDATE_NAME_COLS  = ["이름", "성명", "Name"]

# 폰트 (맥/윈도우 중 하나로 자동 시도)
PREFERRED_FONTS = ["Apple SD Gothic Neo", "Malgun Gothic"]


# =========================
# Streamlit 기본 설정
# =========================
st.set_page_config(page_title="주차별 학생 신청 과목 조회", layout="wide", initial_sidebar_state="expanded")

# 상단 여백/헤더 겹침 방지
st.markdown("""
<style>
.block-container { padding-top: 1.6rem; }  /* 본문 상단 여백 증가 */
h1, .stMarkdown h1 { margin-top: 0.2rem; } /* 제목 상단 마진 보정 */
</style>
""", unsafe_allow_html=True)

# =========================
# 유틸 함수
# =========================
def normalize_text(s: str) -> str:
    """한글 깨짐 방지용 NFC 정규화."""
    return unicodedata.normalize("NFC", str(s)) if s is not None else ""


def find_col(df_cols, candidates):
    """후보 리스트로 컬럼 자동 매칭(완전 일치 우선 → 부분 일치)."""
    lower_cols = {str(c).lower(): c for c in df_cols}
    # 완전 일치
    for cand in candidates:
        k = cand.lower()
        if k in lower_cols:
            return lower_cols[k]
    # 부분 일치
    for cand in candidates:
        pat = re.compile(re.escape(cand), re.IGNORECASE)
        for c in df_cols:
            if pat.search(str(c)):
                return c
    return None


def is_one(x):
    """셀 값이 1(문자 '1' 포함)인지 판정."""
    if pd.isna(x):
        return False
    try:
        return float(str(x).strip()) == 1.0
    except Exception:
        return False


def read_excel_row_values_for_headers(xlsx_path, sheet_name, usecols_range, target_row, headers):
    """
    A:AG 구간의 target_row 값을 읽어 df.columns와 1:1 매핑.
    숫자 변환 실패/공백은 0 처리.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[sheet_name]] if isinstance(sheet_name, int) else (
        wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    )

    start_letter, end_letter = usecols_range.split(":")
    start_idx = column_index_from_string(start_letter)
    end_idx   = column_index_from_string(end_letter)

    vals = []
    for col_idx in range(start_idx, end_idx + 1):
        v = ws.cell(row=target_row, column=col_idx).value
        try:
            v_num = float(str(v).strip())
        except Exception:
            v_num = 0.0
        vals.append(v_num)

    n = min(len(headers), len(vals))
    return {headers[i]: vals[i] for i in range(n)}


def choose_font_name() -> str:
    """사용 가능한 한글 폰트명을 하나 선택."""
    # 엑셀 파일 자체에 폰트가 포함되지는 않지만, 이름을 지정해두면 OS에서 매칭을 시도함.
    for name in PREFERRED_FONTS:
        return name  # 간단히 첫 후보 사용(필요 시 환경 감지 로직 추가 가능)
    return "Arial"


def build_formatted_excel(df_view: pd.DataFrame, file_choice: str, filter_col: str) -> io.BytesIO:
    """
    - 1행: '{파일명(확장자제외)} {filter_col}' 제목(굵게/가운데/병합)
    - 3행: 헤더 [순번, 반, 번호, 이름, 월, 화, 수, 목, 금]
    - 4행~: 데이터 (순번 1부터), 월~금은 빈칸
    - 테두리: 헤더~마지막 데이터 행 모두 얇은 선
    - 정렬: 가운데 정렬(이름도 가운데, 원하면 변경 가능)
    """
    title_text = f"{file_choice.replace('.xlsx','')} {filter_col}".strip()
    title = normalize_text(title_text)

    headers = ["순번", "반", "번호", "이름", "월", "화", "수", "목", "금"]
    n_cols = len(headers)
    font_name = choose_font_name()

    wb = Workbook()
    ws = wb.active
    ws.title = "필터결과"

    # 제목(1행)
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws["A1"].font = Font(name=font_name, size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 6  # 여백

    # 헤더(3행)
    header_row = 3
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=idx, value=h)
        cell.font = Font(name=font_name, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 데이터(4행~)
    start_row = header_row + 1
    for i, (_, row) in enumerate(df_view.iterrows(), start=1):
        r = start_row + i - 1
        ws.cell(row=r, column=1, value=i).font = Font(name=font_name)          # 순번
        ws.cell(row=r, column=2, value=row.iloc[0]).font = Font(name=font_name) # 반
        ws.cell(row=r, column=3, value=row.iloc[1]).font = Font(name=font_name) # 번호
        ws.cell(row=r, column=4, value=row.iloc[2]).font = Font(name=font_name) # 이름
        for c in range(5, 10):
            ws.cell(row=r, column=c).font = Font(name=font_name)                # 월~금 비워둠

    # 테두리/정렬
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    last_row = start_row + len(df_view) - 1 if len(df_view) > 0 else header_row
    end_row = max(header_row, last_row)

    for r in range(header_row, end_row + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 열 너비/행 높이
    widths = {1:6, 2:6, 3:6, 4:16, 5:6, 6:6, 7:6, 8:6, 9:6}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    for r in range(header_row, end_row + 1):
        ws.row_dimensions[r].height = 20

    # 메모리로 저장
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_pdf(df_result: pd.DataFrame, file_choice: str, filter_col: str) -> io.BytesIO:
    """간단한 표 형태 PDF."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = []
    title = Paragraph(f"필터 결과 ({normalize_text(filter_col)} == 1) — {normalize_text(file_choice)}", styles["Title"])
    story.append(title)
    story.append(Spacer(1, 12))

    if df_result.empty:
        story.append(Paragraph("표시할 데이터가 없습니다.", styles["Normal"]))
    else:
        data = [list(df_result.columns)] + df_result.values.tolist()
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
            ("TEXTCOLOR", (0,0), (-1,0), colors.black),
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("GRID", (0,0), (-1,-1), 0.3, colors.grey),
            ("FONTSIZE", (0,0), (-1,-1), 9),
            ("BOTTOMPADDING", (0,0), (-1,0), 6),
        ]))
        story.append(table)

    story.append(Spacer(1, 12))
    story.append(Paragraph(datetime.now().strftime("%Y-%m-%d %H:%M"), styles["Normal"]))
    doc.build(story)
    buf.seek(0)
    return buf


# =========================
# 사이드바: 파일 선택
# =========================
os.makedirs(DATA_DIR, exist_ok=True)
excel_files = [f for f in os.listdir(DATA_DIR) if f.lower().endswith((".xlsx", ".xls"))]
excel_files_sorted = sorted(excel_files)

with st.sidebar:
    st.header("선택")
    if len(excel_files_sorted) == 0:
        st.error("`data/` 폴더에 주차별 엑셀을 넣어주세요. 예: 10월3주차.xlsx, 10월4주차.xlsx")
        st.stop()
    file_choice = st.selectbox("① 주차별 선택", excel_files_sorted, index=0)

selected_path = os.path.join(DATA_DIR, file_choice)

# =========================
# 데이터 로드
# =========================
try:
    df = pd.read_excel(
        selected_path,
        sheet_name=SHEET_NAME,
        header=HEADER_ROW_INDEX,
        usecols=USECOLS_RANGE,
        engine="openpyxl",
    )
except Exception as e:
    st.error(f"엑셀 로드 실패: {e}")
    st.stop()

df.columns = [str(c).strip() for c in df.columns]

# =========================
# 필터 기준 열 후보 만들기
#  - 반/번호/이름 제외
#  - 174행 합계가 0인 컬럼 제외
# =========================
class_col_detected = find_col(df.columns, CANDIDATE_CLASS_COLS)
no_col_detected    = find_col(df.columns, CANDIDATE_NO_COLS)
name_col_detected  = find_col(df.columns, CANDIDATE_NAME_COLS)
blocked_cols = {c for c in [class_col_detected, no_col_detected, name_col_detected] if c is not None}

sum_map_174 = read_excel_row_values_for_headers(
    xlsx_path=selected_path,
    sheet_name=SHEET_NAME,
    usecols_range=USECOLS_RANGE,
    target_row=TARGET_SUM_ROW,
    headers=list(df.columns),
)
zero_sum_cols = {col for col, v in sum_map_174.items() if v == 0.0}

filterable_cols = [c for c in df.columns if c not in blocked_cols and c not in zero_sum_cols]
if len(filterable_cols) == 0:
    filterable_cols = list(df.columns)
    st.warning("필터 가능한 컬럼이 없습니다. (반/번호/이름 제외 + 174행 합계==0 제외) — 임시로 전체 컬럼을 노출합니다. 174행 합계/헤더 확인 필요.")

# =========================
# 사이드바: 필터 기준 열 선택
# =========================
with st.sidebar:
    filter_col = st.selectbox("② 과목 선택", options=filterable_cols)
    st.info(f"**{normalize_text(filter_col)}** 선택됨")

# =========================
# 제목(선택 이후 렌더링)
# =========================
title_text = f"{normalize_text(file_choice).replace('.xlsx','')} {normalize_text(filter_col)} 수강자 명단"
st.markdown(f"<h1 style='text-align:center;'>{title_text}</h1>", unsafe_allow_html=True)

# =========================
# 필터 및 표시
# =========================
mask = df[filter_col].apply(is_one)
df_one = df[mask].copy()

class_col = find_col(df_one.columns, CANDIDATE_CLASS_COLS) or "반"
no_col    = find_col(df_one.columns, CANDIDATE_NO_COLS)    or "번호"
name_col  = find_col(df_one.columns, CANDIDATE_NAME_COLS)  or "이름"

display_cols = [c for c in [class_col, no_col, name_col] if c in df_one.columns]

st.markdown("""
<style>
/* st.dataframe 내 thead/tbody 가운데 정렬(폴백) */
thead tr th { text-align: center !important; }
tbody tr td { text-align: center !important; }
</style>
""", unsafe_allow_html=True)

#st.subheader("결과 미리보기")
if not display_cols:
    st.warning("표시할 컬럼(반/번호/이름)을 찾지 못했습니다. 엑셀 헤더를 확인해주세요.")
    st.dataframe(df_one.head(20), use_container_width=True, height=320)
    df_view = df_one.copy()
else:
    df_view = df_one[display_cols].copy()
    # 행 개수에 따라 높이 자동 조절 (최대 900px)
    row_px = 37   # 행당 평균 픽셀(헤더 제외)
    base  = 80    # 헤더/패딩 여유
    max_h = 900   # 최대 높이
    calc_h = min(max_h, base + row_px * max(5, len(df_view)))
    # ---- 화면 표시용 DataFrame: 순번 열 추가 ----
    df_display = df_view.reset_index(drop=True).copy()
    df_display.insert(0, "순번", range(1, len(df_display) + 1))

    # ---- 표 높이 동적 조절(최대 900px) ----
    row_px = 37   # 행당 픽셀(대략값)
    base   = 100  # 헤더/패딩 여유
    max_h  = 900
    calc_h = min(max_h, base + row_px * max(5, len(df_display)))

    st.dataframe(df_display, use_container_width=True, height=calc_h, hide_index=True)

# =========================
# 다운로드(엑셀 .xlsx만)
# =========================
col1 = st.columns(1)[0]

with col1:
    xlsx_bytes = build_formatted_excel(df_view, file_choice, filter_col)
    st.download_button(
        "엑셀 다운로드",
        data=xlsx_bytes,
        file_name=f"{normalize_text(file_choice).replace('.xlsx','')}_{normalize_text(filter_col)}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

st.success("사이드바에서 주차/과목 선택 → 메인에서 미리보기 → 엑셀(출석부) 저장 가능합니다.")